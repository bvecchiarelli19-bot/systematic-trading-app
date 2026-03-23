"""Trade Journal API — CRUD for journal entries + performance stats."""
import io
import math
from datetime import datetime

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional

from backend.data.database import SessionLocal, JournalEntry

router = APIRouter(prefix="/api/journal", tags=["journal"])


# ── Pydantic Schemas ─────────────────────────────────

class TradeEntryCreate(BaseModel):
    ticker: str
    name: Optional[str] = None
    sector: Optional[str] = None
    entry_price: float
    entry_date: str
    trade_type: str = "equity"
    position_dollars: float = 0
    regime_at_entry: Optional[str] = None
    composite_score: Optional[int] = None
    position_pct: Optional[int] = None
    thesis: Optional[str] = None
    target_price: Optional[float] = None
    stop_price: Optional[float] = None
    asymmetry_ratio: Optional[float] = None
    bias_score: Optional[int] = None
    regime_confirmed: Optional[bool] = None
    vol_percentile: Optional[float] = None
    sma_trend: Optional[str] = None
    hurst: Optional[float] = None
    tail_risk: Optional[float] = None
    predicted_probability: Optional[float] = None


class TradeCloseRequest(BaseModel):
    exit_price: float
    outcome: str  # WIN or LOSS
    regime_correct: bool
    thesis_correct: bool
    key_learning: Optional[str] = ""
    actual_outcome_binary: Optional[int] = None  # for prediction markets


# ── Helpers ───────────────────────────────────────────

def entry_to_dict(e: JournalEntry) -> dict:
    return {
        "id": e.id,
        "ticker": e.ticker,
        "name": e.name,
        "sector": e.sector,
        "entry_price": e.entry_price,
        "entry_date": e.entry_date,
        "trade_type": e.trade_type,
        "position_dollars": e.position_dollars,
        "regime_at_entry": e.regime_at_entry,
        "composite_score": e.composite_score,
        "position_pct": e.position_pct,
        "thesis": e.thesis,
        "target_price": e.target_price,
        "stop_price": e.stop_price,
        "asymmetry_ratio": e.asymmetry_ratio,
        "bias_score": e.bias_score,
        "regime_confirmed": e.regime_confirmed,
        "vol_percentile": e.vol_percentile,
        "sma_trend": e.sma_trend,
        "hurst": e.hurst,
        "tail_risk": e.tail_risk,
        "status": e.status,
        "exit_price": e.exit_price,
        "exit_date": e.exit_date,
        "outcome": e.outcome,
        "regime_correct": e.regime_correct,
        "thesis_correct": e.thesis_correct,
        "key_learning": e.key_learning,
        "pnl_dollars": e.pnl_dollars,
        "pnl_pct": e.pnl_pct,
        "predicted_probability": e.predicted_probability,
        "actual_outcome_binary": e.actual_outcome_binary,
        "created_at": e.created_at,
        "closed_at": e.closed_at,
    }


# ── Endpoints ─────────────────────────────────────────

@router.get("")
async def list_trades(status: str = "all"):
    """List all journal entries, optionally filtered by status."""
    session = SessionLocal()
    try:
        q = session.query(JournalEntry)
        if status == "open":
            q = q.filter(JournalEntry.status == "OPEN")
        elif status == "closed":
            q = q.filter(JournalEntry.status == "CLOSED")
        entries = q.order_by(JournalEntry.id.desc()).all()
        return [entry_to_dict(e) for e in entries]
    finally:
        session.close()


@router.post("")
async def create_trade(data: TradeEntryCreate):
    """Create a new journal entry."""
    session = SessionLocal()
    try:
        entry = JournalEntry(
            ticker=data.ticker,
            name=data.name,
            sector=data.sector,
            entry_price=data.entry_price,
            entry_date=data.entry_date,
            trade_type=data.trade_type,
            position_dollars=data.position_dollars,
            regime_at_entry=data.regime_at_entry,
            composite_score=data.composite_score,
            position_pct=data.position_pct,
            thesis=data.thesis,
            target_price=data.target_price,
            stop_price=data.stop_price,
            asymmetry_ratio=data.asymmetry_ratio,
            bias_score=data.bias_score,
            regime_confirmed=data.regime_confirmed,
            vol_percentile=data.vol_percentile,
            sma_trend=data.sma_trend,
            hurst=data.hurst,
            tail_risk=data.tail_risk,
            predicted_probability=data.predicted_probability,
            status="OPEN",
            created_at=datetime.utcnow().isoformat(),
        )
        session.add(entry)
        session.commit()
        session.refresh(entry)
        return entry_to_dict(entry)
    finally:
        session.close()


@router.post("/{trade_id}/close")
async def close_trade(trade_id: int, data: TradeCloseRequest):
    """Close a trade and record outcome."""
    session = SessionLocal()
    try:
        entry = session.query(JournalEntry).filter_by(id=trade_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="Trade not found")
        if entry.status == "CLOSED":
            raise HTTPException(status_code=400, detail="Trade already closed")

        entry.exit_price = data.exit_price
        entry.outcome = data.outcome.upper()
        entry.regime_correct = data.regime_correct
        entry.thesis_correct = data.thesis_correct
        entry.key_learning = data.key_learning
        entry.status = "CLOSED"
        entry.exit_date = datetime.utcnow().strftime("%Y-%m-%d")
        entry.closed_at = datetime.utcnow().isoformat()

        # Calculate P&L
        if entry.entry_price and entry.entry_price > 0:
            entry.pnl_pct = round((data.exit_price - entry.entry_price) / entry.entry_price * 100, 2)
            if entry.position_dollars and entry.position_dollars > 0:
                shares = entry.position_dollars / entry.entry_price
                entry.pnl_dollars = round(shares * (data.exit_price - entry.entry_price), 2)

        # Prediction market Brier score data
        if data.actual_outcome_binary is not None:
            entry.actual_outcome_binary = data.actual_outcome_binary

        session.commit()
        session.refresh(entry)
        return entry_to_dict(entry)
    finally:
        session.close()


@router.delete("/{trade_id}")
async def delete_trade(trade_id: int):
    """Delete a journal entry."""
    session = SessionLocal()
    try:
        entry = session.query(JournalEntry).filter_by(id=trade_id).first()
        if not entry:
            raise HTTPException(status_code=404, detail="Trade not found")
        session.delete(entry)
        session.commit()
        return {"deleted": True, "id": trade_id}
    finally:
        session.close()


@router.get("/stats")
async def get_stats():
    """Performance dashboard statistics."""
    session = SessionLocal()
    try:
        all_trades = session.query(JournalEntry).all()
        closed = [t for t in all_trades if t.status == "CLOSED"]
        open_trades = [t for t in all_trades if t.status == "OPEN"]

        total = len(all_trades)
        total_closed = len(closed)
        wins = [t for t in closed if t.outcome == "WIN"]
        losses = [t for t in closed if t.outcome == "LOSS"]
        win_rate = (len(wins) / total_closed * 100) if total_closed > 0 else 0

        # Regime accuracy
        regime_assessed = [t for t in closed if t.regime_correct is not None]
        regime_correct_count = len([t for t in regime_assessed if t.regime_correct])
        regime_accuracy = (regime_correct_count / len(regime_assessed) * 100) if regime_assessed else 0

        # Thesis accuracy
        thesis_assessed = [t for t in closed if t.thesis_correct is not None]
        thesis_correct_count = len([t for t in thesis_assessed if t.thesis_correct])
        thesis_accuracy = (thesis_correct_count / len(thesis_assessed) * 100) if thesis_assessed else 0

        # Avg asymmetry
        asym_trades = [t for t in all_trades if t.asymmetry_ratio and t.asymmetry_ratio > 0]
        avg_asymmetry = (sum(t.asymmetry_ratio for t in asym_trades) / len(asym_trades)) if asym_trades else 0

        # P&L
        total_pnl_dollars = sum(t.pnl_dollars or 0 for t in closed)
        pnl_values = [t.pnl_pct for t in closed if t.pnl_pct is not None]
        avg_pnl_pct = (sum(pnl_values) / len(pnl_values)) if pnl_values else 0

        # Drawdown (peak-to-trough from cumulative P&L)
        cumulative = 0
        peak = 0
        max_dd = 0
        for t in sorted(closed, key=lambda x: x.closed_at or ""):
            cumulative += (t.pnl_dollars or 0)
            if cumulative > peak:
                peak = cumulative
            dd = peak - cumulative
            if dd > max_dd:
                max_dd = dd

        # Brier score for prediction market trades
        prediction_trades = [t for t in closed if t.trade_type == "prediction"
                             and t.predicted_probability is not None
                             and t.actual_outcome_binary is not None]
        brier_score = None
        if prediction_trades:
            brier_sum = sum(
                (t.predicted_probability - t.actual_outcome_binary) ** 2
                for t in prediction_trades
            )
            brier_score = round(brier_sum / len(prediction_trades), 4)

        return {
            "total_trades": total,
            "open_trades": len(open_trades),
            "closed_trades": total_closed,
            "wins": len(wins),
            "losses": len(losses),
            "win_rate": round(win_rate, 1),
            "regime_accuracy": round(regime_accuracy, 1),
            "thesis_accuracy": round(thesis_accuracy, 1),
            "avg_asymmetry": round(avg_asymmetry, 1),
            "total_pnl_dollars": round(total_pnl_dollars, 2),
            "avg_pnl_pct": round(avg_pnl_pct, 2),
            "max_drawdown": round(max_dd, 2),
            "brier_score": brier_score,
            "prediction_trade_count": len(prediction_trades),
        }
    finally:
        session.close()


@router.get("/export")
async def export_xlsx():
    """Export all journal entries as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    session = SessionLocal()
    try:
        entries = session.query(JournalEntry).order_by(JournalEntry.id.asc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trade Journal"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a2233", end_color="1a2233", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="2a3a50"),
            right=Side(style="thin", color="2a3a50"),
            top=Side(style="thin", color="2a3a50"),
            bottom=Side(style="thin", color="2a3a50"),
        )

        headers = [
            "ID", "Ticker", "Name", "Sector", "Type", "Status",
            "Entry Date", "Entry Price", "Position $", "Regime", "Composite",
            "Vol %ile", "SMA", "Hurst", "CVaR",
            "Thesis", "Target", "Stop", "Asymmetry",
            "Bias Score", "Regime Confirmed",
            "Exit Date", "Exit Price", "Outcome", "P&L $", "P&L %",
            "Regime Correct", "Thesis Correct", "Key Learning",
            "Pred Prob", "Actual Outcome",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, e in enumerate(entries, 2):
            values = [
                e.id, e.ticker, e.name, e.sector, e.trade_type, e.status,
                e.entry_date, e.entry_price, e.position_dollars, e.regime_at_entry, e.composite_score,
                e.vol_percentile, e.sma_trend, e.hurst, e.tail_risk,
                e.thesis, e.target_price, e.stop_price, e.asymmetry_ratio,
                e.bias_score, e.regime_confirmed,
                e.exit_date, e.exit_price, e.outcome, e.pnl_dollars, e.pnl_pct,
                e.regime_correct, e.thesis_correct, e.key_learning,
                e.predicted_probability, e.actual_outcome_binary,
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 40))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

        # Write to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"trade_journal_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        session.close()
